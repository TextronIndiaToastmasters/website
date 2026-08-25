"""Converts the club's Attendance/Role-Takers Excel workbook into
data/stats.json for the website. All months/years live together in a single
workbook, excel-data/stats.xlsm — new meetings are just added as extra
columns/rows. A GitHub Actions workflow (.github/workflows/generate-stats.yml)
runs this automatically and commits the regenerated JSON whenever
excel-data/stats.xlsm changes on the main branch, so GitHub Pages picks it up
without any manual step. You can still run it locally to preview changes.

Usage:
    python scripts/generate_stats.py

Requires: pip install openpyxl
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "excel-data" / "stats.xlsm"
OUT = ROOT / "data" / "stats.json"


def clean_name(value):
    return " ".join(str(value).split())


def normalize(text):
    """Collapses whitespace and lowercases, so sheet/column names match
    regardless of stray spaces or letter casing."""
    return " ".join(str(text).split()).lower()


def get_sheet(wb, *accepted_names):
    """Finds a worksheet by name, tolerant of case and extra whitespace, and
    accepting any of the given alias spellings (e.g. Attendence/Attendance)."""
    targets = {normalize(n) for n in accepted_names}
    for sheet_name in wb.sheetnames:
        if normalize(sheet_name) in targets:
            return wb[sheet_name]
    raise KeyError(
        f"No sheet matching {accepted_names!r} found. Available sheets: {wb.sheetnames}"
    )


def find_col(header_row, label):
    target = normalize(label)
    for cell in header_row:
        if isinstance(cell.value, str) and normalize(cell.value) == target:
            return cell.column
    return None


def date_columns(header_row):
    cols = []
    for cell in header_row:
        if cell.column == 1:
            continue
        if isinstance(cell.value, datetime):
            cols.append(cell.column)
        elif cols:
            break
    return cols


def build_attendance(wb):
    ws = get_sheet(wb, "Attendence", "Attendance")
    header = ws[1]
    date_cols = date_columns(header)
    dates = [ws.cell(row=1, column=c).value.strftime("%Y-%m-%d") for c in date_cols]

    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name = row[0].value
        if not name:
            continue
        attended = sum(1 for c in date_cols if row[c - 1].value)
        total = len(date_cols)
        points = attended * 2
        percentage = round((attended / total) * 100, 1) if total else 0
        entries.append({
            "name": clean_name(name),
            "meetingsAttended": attended,
            "totalMeetings": total,
            "points": points,
            "percentage": percentage,
        })

    entries.sort(key=lambda e: e["points"], reverse=True)
    return dates, entries


def build_role_takers(wb):
    ws = get_sheet(wb, "Role Takers (points)")
    header = ws[1]
    total_rt_col = find_col(header, "Total RT")
    total_combined_col = find_col(header, "Total Attendance & RT (No Need)")

    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name = row[0].value
        if not name:
            continue
        entries.append({
            "name": clean_name(name),
            "roleTakerPoints": (row[total_rt_col - 1].value if total_rt_col else 0) or 0,
            "totalCombinedPoints": (row[total_combined_col - 1].value if total_combined_col else 0) or 0,
        })

    entries.sort(key=lambda e: e["roleTakerPoints"], reverse=True)
    return entries


def build_buddy_olympics(wb):
    ws = get_sheet(wb, "Buddy Olympics")
    header = ws[1]
    date_cols = set(date_columns(header))
    fixed_cols = {1}
    for label in ("Total", "Total Attendance & RT", "Total - A"):
        col = find_col(header, label)
        if col:
            fixed_cols.add(col)

    team_cols = [
        (cell.column, cell.value)
        for cell in header
        if cell.column not in date_cols and cell.column not in fixed_cols and cell.value
    ]

    totals_row = ws[2]
    entries = [
        {"team": clean_name(name), "points": totals_row[col - 1].value or 0}
        for col, name in team_cols
    ]
    entries.sort(key=lambda e: e["points"], reverse=True)
    return entries


def build_people_choice(wb):
    ws = get_sheet(wb, "People Choice Award")
    header = ws[1]
    dates = [
        (cell.column, cell.value.strftime("%Y-%m-%d"))
        for cell in header
        if cell.column != 1 and isinstance(cell.value, datetime)
    ]

    by_date = {d: {} for _, d in dates}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        category = row[0].value
        if not category:
            continue
        for col, d in dates:
            winner = row[col - 1].value
            if winner:
                by_date[d][category] = clean_name(winner)

    meetings = [{"date": d, "awards": awards} for d, awards in by_date.items() if awards]
    meetings.sort(key=lambda m: m["date"], reverse=True)
    return meetings


def main():
    if not WORKBOOK.exists():
        print(f"Workbook not found: {WORKBOOK}", file=sys.stderr)
        sys.exit(1)

    print(f"Processing {WORKBOOK.name}...")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    dates, attendance = build_attendance(wb)

    stats = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "meetingDates": dates,
        "attendance": attendance,
        "roleTakers": build_role_takers(wb),
        "buddyOlympics": build_buddy_olympics(wb),
        "peopleChoiceAwards": build_people_choice(wb),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(stats, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
